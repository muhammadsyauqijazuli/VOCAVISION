import io
from collections import Counter
from datetime import datetime
from typing import Any

from flask import Blueprint, request, jsonify, make_response
from flask_jwt_extended import jwt_required
from ..models import Student, Prediction, SHAPAnalysis, Intervention, User

from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

export_bp = Blueprint('export', __name__)


EXCEL_COLUMNS = [
    ('nama_siswa', 'Nama Siswa'),
    ('nisn', 'NISN'),
    ('jam_belajar_per_hari', 'Jam Belajar / Hari'),
    ('presentase_kehadiran', 'Presentase Kehadiran'),
    ('nilai_rata_rata_raport', 'Nilai Rata-rata Raport'),
    ('skor_time_management', 'Skor Time Management'),
    ('jam_tidur', 'Jam Tidur'),
    ('screen_time', 'Screen Time'),
    ('kehadiran_pelatihan_industry', 'Kehadiran Pelatihan Industry'),
    ('motivasi_akademik', 'Motivasi Akademik'),
    ('exam_score', 'Nilai Ujian Aktual'),
    ('gender', 'Gender'),
    ('rata_rata_pemasukan_keluarga', 'Rata-rata Pemasukan Keluarga'),
    ('pendidikan_terakhir_orang_tua', 'Pendidikan Terakhir Orang Tua'),
    ('kerja_sampingan', 'Kerja Sampingan'),
    ('study_environment', 'Study Environment'),
    ('kompetensi_skill_level', 'Kompetensi Skill Level'),
    ('industry_readiness', 'Industry Readiness'),
    ('stress_level', 'Stress Level'),
    ('predicted_score', 'Predicted Exam Score'),
    ('risk_status', 'Risk Status'),
    ('model_version', 'Model Version'),
    ('predicted_at', 'Predicted At'),
]


def _format_value(value):
    if value is None:
        return ''
    if isinstance(value, float):
        return round(value, 2)
    return value


def _format_datetime(dt):
    if not dt:
        return ''
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d %H:%M')
    return str(dt)


def build_latest_prediction_map():
    latest_by_student_id = {}

    for prediction in Prediction.query.all():
        current = latest_by_student_id.get(prediction.student_id)
        if current is None:
            latest_by_student_id[prediction.student_id] = prediction
            continue

        current_time = current.created_at
        new_time = prediction.created_at
        if current_time is None or (new_time is not None and new_time > current_time):
            latest_by_student_id[prediction.student_id] = prediction

    return latest_by_student_id


def gather_rows(class_id=None, risk_status=None):
    # class_id is reserved for future classroom filtering.
    students = Student.query.order_by(Student.nama_siswa.asc()).all()
    latest_by_student_id = build_latest_prediction_map()
    rows = []

    for student in students:
        prediction = latest_by_student_id.get(student.id)
        row = {
            'nama_siswa': student.nama_siswa,
            'nisn': student.nisn,
            'jam_belajar_per_hari': student.jam_belajar_per_hari,
            'presentase_kehadiran': student.presentase_kehadiran,
            'nilai_rata_rata_raport': student.nilai_rata_rata_raport,
            'skor_time_management': student.skor_time_management,
            'jam_tidur': student.jam_tidur,
            'screen_time': student.screen_time,
            'kehadiran_pelatihan_industry': student.kehadiran_pelatihan_industry,
            'motivasi_akademik': student.motivasi_akademik,
            'exam_score': student.exam_score,
            'gender': student.gender,
            'rata_rata_pemasukan_keluarga': student.rata_rata_pemasukan_keluarga,
            'pendidikan_terakhir_orang_tua': student.pendidikan_terakhir_orang_tua,
            'kerja_sampingan': student.kerja_sampingan,
            'study_environment': student.study_environment,
            'kompetensi_skill_level': student.kompetensi_skill_level,
            'industry_readiness': student.industry_readiness,
            'stress_level': student.stress_level,
            'predicted_score': prediction.predicted_exam_score if prediction else None,
            'risk_status': prediction.risk_status if prediction else None,
            'model_version': prediction.model_version if prediction else '',
            'predicted_at': prediction.created_at if prediction else None,
        }

        if risk_status and row['risk_status'] != risk_status:
            continue

        rows.append(row)

    return rows


def _draw_vocavision_header_footer(canvas, doc):
    page_width, page_height = landscape(A4)
    page_number = canvas.getPageNumber()

    canvas.saveState()

    # Header band
    canvas.setFillColor(colors.HexColor('#0F766E'))
    canvas.rect(0, page_height - 18 * mm, page_width, 18 * mm, fill=1, stroke=0)

    canvas.setFillColor(colors.white)
    canvas.setFont('Helvetica-Bold', 12)
    canvas.drawString(16 * mm, page_height - 12 * mm, 'VOCAVISION')
    canvas.setFont('Helvetica', 9)
    canvas.drawRightString(page_width - 16 * mm, page_height - 12 * mm, 'Report Akademik Per Siswa')

    # Footer
    canvas.setStrokeColor(colors.HexColor('#CBD5E1'))
    canvas.setLineWidth(0.5)
    canvas.line(16 * mm, 14 * mm, page_width - 16 * mm, 14 * mm)

    canvas.setFillColor(colors.HexColor('#334155'))
    canvas.setFont('Helvetica', 8)
    canvas.drawString(16 * mm, 9 * mm, 'VOCAVISION | Vocational Student Predictive Analytics')
    canvas.drawRightString(page_width - 16 * mm, 9 * mm, f'Halaman {page_number}')

    canvas.restoreState()


def _build_risk_pie_chart(rows):
    counts = Counter(row.get('risk_status') or 'Belum Ada Prediksi' for row in rows)
    labels = ['Rendah', 'Netral', 'Tinggi', 'Belum Ada Prediksi']
    values = [counts.get(label, 0) for label in labels]

    drawing = Drawing(250, 220)
    chart = Pie()
    chart.x = 18
    chart.y = 20
    chart.width = 118
    chart.height = 118
    chart.labels = [f'{label} ({values[index]})' for index, label in enumerate(labels)]
    chart.data = values
    chart.sideLabels = True
    chart.slices.strokeWidth = 0.5
    chart.slices[0].fillColor = colors.HexColor('#E74C3C')
    chart.slices[1].fillColor = colors.HexColor('#F39C12')
    chart.slices[2].fillColor = colors.HexColor('#A3E4D7')
    chart.slices[3].fillColor = colors.HexColor('#94A3B8')

    drawing.add(chart)
    drawing.add(String(150, 185, 'Distribusi Status Risiko', fontName='Helvetica-Bold', fontSize=10, fillColor=colors.HexColor('#0F172A')))
    drawing.add(String(150, 168, f'Total siswa: {len(rows)}', fontName='Helvetica', fontSize=9, fillColor=colors.HexColor('#334155')))
    return drawing


def _build_stress_bar_chart(rows):
    stress_labels = ['Rendah', 'Sedang', 'Berat']
    grouped = {label: [] for label in stress_labels}

    for row in rows:
        stress_level = row.get('stress_level')
        predicted_score = row.get('predicted_score')
        if stress_level in grouped and isinstance(predicted_score, (int, float)):
            grouped[stress_level].append(float(predicted_score))

    averages = []
    for label in stress_labels:
        values = grouped[label]
        avg = (sum(values) / len(values)) if values else 0
        averages.append(round(avg, 2))

    drawing = Drawing(250, 220)
    chart = VerticalBarChart()
    chart.x = 28
    chart.y = 35
    chart.height = 140
    chart.width = 185
    chart.data = [averages]
    chart.strokeColor = colors.HexColor('#94A3B8')
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = 100
    chart.valueAxis.valueStep = 20
    chart.categoryAxis.categoryNames = stress_labels
    chart.categoryAxis.labels.angle = 0
    chart.categoryAxis.labels.boxAnchor = 'n'
    chart.bars[0].fillColor = colors.HexColor('#0F766E')

    drawing.add(chart)
    drawing.add(String(28, 185, 'Rata-rata Prediksi Skor per Stress Level', fontName='Helvetica-Bold', fontSize=10, fillColor=colors.HexColor('#0F172A')))
    return drawing


def _build_attendance_bar_chart(rows):
    attendance_rows = [
        row for row in rows
        if isinstance(row.get('presentase_kehadiran'), (int, float))
        and isinstance(row.get('predicted_score'), (int, float))
    ]
    attendance_rows.sort(key=lambda item: float(item['presentase_kehadiran']), reverse=True)
    attendance_rows = attendance_rows[:6]

    # Jika tidak ada data, kembalikan placeholder
    if not attendance_rows:
        drawing = Drawing(420, 220)
        drawing.add(String(40, 100, "Data kehadiran tidak tersedia",
                          fontName='Helvetica', fontSize=10,
                          fillColor=colors.HexColor('#94A3B8')))
        return drawing

    averages = [round(float(row['predicted_score']), 2) for row in attendance_rows]
    labels = [(row.get('nama') or '')[:12] for row in attendance_rows]


    drawing = Drawing(420, 220)
    chart = VerticalBarChart()
    chart.x = 40
    chart.y = 35
    chart.height = 140
    chart.width = 320
    chart.data = [averages]
    chart.strokeColor = colors.HexColor('#94A3B8')
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = 100
    chart.valueAxis.valueStep = 20
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.angle = 25
    chart.categoryAxis.labels.boxAnchor = 'n'
    chart.bars[0].fillColor = colors.HexColor('#3BA99C')

    drawing.add(chart)
    drawing.add(String(40, 185, '6 Siswa dengan Kehadiran Tertinggi', fontName='Helvetica-Bold', fontSize=10, fillColor=colors.HexColor('#0F172A')))
    return drawing


def _build_top_risky_students_table(rows):
    risky_rows = [
        row for row in rows
        if row.get('risk_status') in {'Rendah', 'Netral'} and isinstance(row.get('predicted_score'), (int, float))
    ]
    risky_rows.sort(key=lambda item: float(item['predicted_score']))

    table_rows = [['Nama', 'NISN', 'Prediksi Skor', 'Status Risiko', 'Stress Level']]
    for row in risky_rows[:5]:
        table_rows.append([
            _format_value(row.get('nama_siswa')),
            _format_value(row.get('nisn')),
            _format_value(row.get('predicted_score')),
            _format_value(row.get('risk_status')),
            _format_value(row.get('stress_level')),
        ])

    if len(table_rows) == 1:
        table_rows.append(['-', '-', '-', '-', '-'])

    return table_rows


def _build_metric_card(title, value, caption, accent_color):
    return Table([
        [Paragraph(f'<font color="#64748B" size="8"><b>{title}</b></font>', getSampleStyleSheet()['BodyText'])],
        [Paragraph(f'<font color="#0F172A" size="20"><b>{value}</b></font>', getSampleStyleSheet()['BodyText'])],
        [Paragraph(f'<font color="#64748B" size="8">{caption}</font>', getSampleStyleSheet()['BodyText'])],
    ], colWidths=[110])


def _build_dashboard_header(title_text, subtitle_text):
    title_style = ParagraphStyle(
        name='DashboardTitle',
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.white,
        spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        name='DashboardSubtitle',
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#DCFCE7'),
    )
    return [
        Paragraph(title_text, title_style),
        Spacer(1, 3),
        Paragraph(subtitle_text, subtitle_style),
    ]


def _build_table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#F8FAFC')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ])


def _slugify_filename(value: str) -> str:
    normalized = ''.join(character if character.isalnum() else '_' for character in (value or '').strip().lower())
    while '__' in normalized:
        normalized = normalized.replace('__', '_')
    return normalized.strip('_') or 'siswa'


def _latest_prediction_for_student(student_id: str):
    return Prediction.query.filter_by(student_id=student_id).order_by(Prediction.created_at.desc()).first()


def _shap_rows_for_prediction(prediction_id: str):
    shap_rows = SHAPAnalysis.query.filter_by(prediction_id=prediction_id).all()
    shap_rows.sort(key=lambda row: abs(float(row.impact_value or 0)), reverse=True)
    return shap_rows


def _interventions_for_student(student_id: str):
    return Intervention.query.filter_by(student_id=student_id).order_by(Intervention.action_date.desc()).all()


def _build_student_report(student_id: str):
    student = Student.query.get(student_id)
    if not student:
        return None, ('Siswa tidak ditemukan', 404)

    prediction = _latest_prediction_for_student(student_id)
    if not prediction:
        return None, ('Prediksi tidak ditemukan', 404)

    shap_rows = _shap_rows_for_prediction(prediction.id)
    interventions = _interventions_for_student(student_id)

    guru_names: dict[str, str] = {}
    for intervention in interventions:
        if intervention.guru_id and intervention.guru_id not in guru_names:
            guru = User.query.get(intervention.guru_id)
            guru_names[intervention.guru_id] = guru.nama if guru else '-'

    report = {
        'student': student,
        'prediction': prediction,
        'shap_rows': shap_rows,
        'interventions': interventions,
        'guru_names': guru_names,
        'generated_at': datetime.now(),
    }

    return report, None


def _student_summary_rows(report: dict[str, Any]):
    student = report['student']
    prediction = report['prediction']
    return [
        ('Profil Siswa', student.nama_siswa),
        ('NISN', student.nisn),
        ('Skor Prediksi', _format_value(prediction.predicted_exam_score)),
        ('Status Risiko', prediction.risk_status),
        ('Versi Model', prediction.model_version),
        ('Tanggal Prediksi', _format_datetime(prediction.created_at)),
        ('Waktu Laporan', _format_datetime(report['generated_at'])),
        ('Jumlah Faktor SHAP', len(report['shap_rows'])),
        ('Jumlah Intervensi', len(report['interventions'])),
    ]


def _student_variable_rows(report: dict[str, Any]):
    student = report['student']
    prediction = report['prediction']
    return [
        ['Field', 'Value'],
        ['Profil Siswa', _format_value(student.nama_siswa)],
        ['NISN', _format_value(student.nisn)],
        ['Jam Belajar / Hari', _format_value(student.jam_belajar_per_hari)],
        ['Presentase Kehadiran', _format_value(student.presentase_kehadiran)],
        ['Nilai Rata-rata Raport', _format_value(student.nilai_rata_rata_raport)],
        ['Skor Time Management', _format_value(student.skor_time_management)],
        ['Jam Tidur', _format_value(student.jam_tidur)],
        ['Screen Time', _format_value(student.screen_time)],
        ['Kehadiran Pelatihan Industry', _format_value(student.kehadiran_pelatihan_industry)],
        ['Motivasi Akademik', _format_value(student.motivasi_akademik)],
        ['Exam Score Aktual', _format_value(student.exam_score)],
        ['Gender', _format_value(student.gender)],
        ['Rata-rata Pemasukan Keluarga', _format_value(student.rata_rata_pemasukan_keluarga)],
        ['Pendidikan Terakhir Orang Tua', _format_value(student.pendidikan_terakhir_orang_tua)],
        ['Kerja Sampingan', _format_value(student.kerja_sampingan)],
        ['Study Environment', _format_value(student.study_environment)],
        ['Kompetensi Skill Level', _format_value(student.kompetensi_skill_level)],
        ['Industry Readiness', _format_value(student.industry_readiness)],
        ['Stress Level', _format_value(student.stress_level)],
        ['Skor Prediksi', _format_value(prediction.predicted_exam_score)],
        ['Status Risiko', _format_value(prediction.risk_status)],
        ['Versi Model', _format_value(prediction.model_version)],
        ['Tanggal Prediksi', _format_datetime(prediction.created_at)],
    ]


def _student_shap_rows(report: dict[str, Any]):
    rows = [['Faktor', 'Dampak', 'Saran']]
    for item in report['shap_rows']:
        rows.append([
            _format_value(item.feature_name),
            _format_value(item.impact_value),
            _format_value(item.suggestion_text),
        ])

    if len(rows) == 1:
        rows.append(['-', '-', 'Belum ada analisis SHAP untuk siswa ini'])

    return rows


def _student_intervention_rows(report: dict[str, Any]):
    rows = [['Tanggal', 'Guru', 'Catatan']]
    for item in report['interventions']:
        rows.append([
            _format_datetime(item.action_date),
            report['guru_names'].get(item.guru_id, '-'),
            _format_value(item.note),
        ])

    if len(rows) == 1:
        rows.append(['-', '-', 'Belum ada intervensi untuk siswa ini'])

    return rows


def _to_float(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _build_student_story(report: dict[str, Any]):
    student = report['student']
    prediction = report['prediction']
    shap_rows = report['shap_rows']

    score = _to_float(prediction.predicted_exam_score)
    attendance = _to_float(student.presentase_kehadiran)
    study_hours = _to_float(student.jam_belajar_per_hari)
    sleep_hours = _to_float(student.jam_tidur)
    screen_time = _to_float(student.screen_time)
    motivation = _to_float(student.motivasi_akademik)
    time_management = _to_float(student.skor_time_management)

    strengths = []
    concerns = []
    actions = []

    if attendance is not None and attendance >= 85:
        strengths.append(f'Kehadiran siswa sudah berada di {attendance:.0f}%, sehingga proses belajar di kelas masih cukup stabil.')
    elif attendance is not None:
        concerns.append(f'Kehadiran siswa masih {attendance:.0f}%, sehingga potensi tertinggal materi perlu dipantau lebih ketat.')
        actions.append('Pantau absensi mingguan dan beri pengingat saat kehadiran mulai turun.')

    if study_hours is not None and study_hours >= 3:
        strengths.append(f'Jam belajar harian berada di {study_hours:.1f} jam, menunjukkan ada usaha belajar yang sudah terbentuk.')
    elif study_hours is not None:
        concerns.append(f'Jam belajar harian baru {study_hours:.1f} jam, masih perlu dorongan agar lebih konsisten.')
        actions.append('Buat target belajar harian yang realistis dan terukur.')

    if sleep_hours is not None and sleep_hours >= 7:
        strengths.append(f'Pola tidur berada di {sleep_hours:.1f} jam per malam, cukup baik untuk menjaga fokus.')
    elif sleep_hours is not None:
        concerns.append(f'Jam tidur hanya {sleep_hours:.1f} jam, yang bisa menurunkan fokus dan stamina belajar.')
        actions.append('Dorong rutinitas tidur 7-8 jam agar konsentrasi lebih stabil.')

    if screen_time is not None and screen_time <= 4:
        strengths.append(f'Screen time masih terkendali di {screen_time:.1f} jam per hari.')
    elif screen_time is not None:
        concerns.append(f'Screen time mencapai {screen_time:.1f} jam per hari, sehingga distraksi belajar perlu dikurangi.')
        actions.append('Batasi penggunaan gawai di luar kebutuhan belajar.')

    if motivation is not None and motivation >= 70:
        strengths.append(f'Motivasi akademik berada di {motivation:.0f}, menandakan dorongan belajar masih cukup bagus.')
    elif motivation is not None:
        concerns.append(f'Motivasi akademik masih {motivation:.0f}, sehingga pendampingan perlu dibuat lebih personal.')
        actions.append('Gunakan target jangka pendek dan umpan balik rutin untuk menjaga semangat belajar.')

    if time_management is not None and time_management >= 70:
        strengths.append(f'Skor manajemen waktu mencapai {time_management:.0f}, cukup mendukung disiplin belajar.')
    elif time_management is not None:
        concerns.append(f'Skor manajemen waktu masih {time_management:.0f}, jadi pengaturan jadwal belajar perlu diperkuat.')
        actions.append('Bantu siswa menyusun jadwal belajar harian yang lebih konsisten.')

    if student.study_environment and student.study_environment.lower() in {'kondusif', 'cukup kondusif'}:
        strengths.append(f'Lingkungan belajar tergolong {student.study_environment.lower()}, relatif mendukung konsentrasi.')
    elif student.study_environment:
        concerns.append(f'Lingkungan belajar masih {student.study_environment.lower()}, yang bisa memengaruhi kualitas fokus.')
        actions.append('Pastikan ruang belajar minim gangguan dan cukup nyaman.')

    if student.stress_level and student.stress_level.lower() == 'rendah':
        strengths.append('Tingkat stres terpantau rendah, sehingga siswa lebih siap menerima beban akademik.')
    elif student.stress_level:
        concerns.append(f'Tingkat stres berada pada level {student.stress_level.lower()}, sehingga perlu pemantauan emosional.')
        actions.append('Lakukan cek kondisi psikologis ringan dan dukungan emosional berkala.')

    if student.kompetensi_skill_level and student.kompetensi_skill_level.lower() == 'tinggi':
        strengths.append('Kompetensi skill berada di level tinggi, ini modal baik untuk pembelajaran lanjutan.')
    elif student.kompetensi_skill_level:
        concerns.append(f'Kompetensi skill masih {student.kompetensi_skill_level.lower()}, perlu penguatan konsep inti.')
        actions.append('Berikan latihan tambahan atau proyek kecil untuk memperkuat skill utama.')

    if student.industry_readiness and student.industry_readiness.lower() == 'siap':
        strengths.append('Kesiapan industri sudah baik, menunjukkan orientasi belajar yang matang.')
    elif student.industry_readiness:
        concerns.append(f'Kesiapan industri masih {student.industry_readiness.lower()}, sehingga pengalaman praktik perlu ditambah.')
        actions.append('Arahkan ke pelatihan praktik atau simulasi berbasis proyek.')

    top_shap_texts = []
    for item in shap_rows[:4]:
        suggestion = (item.suggestion_text or '').strip()
        if suggestion and suggestion not in top_shap_texts:
            top_shap_texts.append(suggestion)

    if not actions:
        actions.append('Pertahankan rutinitas belajar yang sudah berjalan dan lakukan monitoring berkala.')

    summary_parts = []
    if score is not None:
        summary_parts.append(f'Skor prediksi siswa berada di {score:.2f}.')
    if prediction.risk_status:
        summary_parts.append(f'Status risiko terdeteksi sebagai {prediction.risk_status}.')
    if score is not None and score >= 75:
        summary_parts.append('Secara umum siswa berada pada zona aman, tetapi beberapa kebiasaan tetap perlu dijaga agar hasil stabil.')
    elif score is not None and score >= 65:
        summary_parts.append('Siswa berada pada zona waspada, sehingga guru perlu menjaga konsistensi belajar dan kehadiran.')
    else:
        summary_parts.append('Siswa berada pada zona prioritas, sehingga tindak lanjut perlu dilakukan lebih cepat dan lebih personal.')

    return {
        'summary_text': ' '.join(summary_parts),
        'strengths': strengths[:4],
        'concerns': concerns[:4],
        'actions': actions[:5],
        'top_shap_texts': top_shap_texts,
    }


def _student_appendix_rows(report: dict[str, Any]):
    student = report['student']
    prediction = report['prediction']
    return [
        ['Field', 'Value'],
        ['Profil Siswa', _format_value(student.nama_siswa)],
        ['NISN', _format_value(student.nisn)],
        ['Gender', _format_value(student.gender)],
        ['Jam Belajar / Hari', _format_value(student.jam_belajar_per_hari)],
        ['Presentase Kehadiran', _format_value(student.presentase_kehadiran)],
        ['Nilai Rata-rata Raport', _format_value(student.nilai_rata_rata_raport)],
        ['Skor Time Management', _format_value(student.skor_time_management)],
        ['Jam Tidur', _format_value(student.jam_tidur)],
        ['Screen Time', _format_value(student.screen_time)],
        ['Kehadiran Pelatihan Industry', _format_value(student.kehadiran_pelatihan_industry)],
        ['Motivasi Akademik', _format_value(student.motivasi_akademik)],
        ['Exam Score Aktual', _format_value(student.exam_score)],
        ['Rata-rata Pemasukan Keluarga', _format_value(student.rata_rata_pemasukan_keluarga)],
        ['Pendidikan Terakhir Orang Tua', _format_value(student.pendidikan_terakhir_orang_tua)],
        ['Kerja Sampingan', _format_value(student.kerja_sampingan)],
        ['Study Environment', _format_value(student.study_environment)],
        ['Kompetensi Skill Level', _format_value(student.kompetensi_skill_level)],
        ['Industry Readiness', _format_value(student.industry_readiness)],
        ['Stress Level', _format_value(student.stress_level)],
        ['Skor Prediksi', _format_value(prediction.predicted_exam_score)],
        ['Status Risiko', _format_value(prediction.risk_status)],
        ['Versi Model', _format_value(prediction.model_version)],
        ['Tanggal Prediksi', _format_datetime(prediction.created_at)],
    ]


def _latest_predictions_by_student():
    latest = {}
    for prediction in Prediction.query.all():
        current = latest.get(prediction.student_id)
        if current is None:
            latest[prediction.student_id] = prediction
            continue

        current_time = current.created_at
        new_time = prediction.created_at
        if current_time is None or (new_time is not None and new_time > current_time):
            latest[prediction.student_id] = prediction

    return latest


def _analytics_dashboard_rows():
    latest_predictions = _latest_predictions_by_student()
    students = Student.query.order_by(Student.nama_siswa.asc()).all()
    rows = []

    for student in students:
        prediction = latest_predictions.get(student.id)
        rows.append({
            'student_id': student.id,
            'nama': student.nama_siswa,
            'nisn': student.nisn,
            'jam_belajar_per_hari': student.jam_belajar_per_hari,
            'presentase_kehadiran': student.presentase_kehadiran,
            'nilai_rata_rata_raport': student.nilai_rata_rata_raport,
            'skor_time_management': student.skor_time_management,
            'jam_tidur': student.jam_tidur,
            'screen_time': student.screen_time,
            'kehadiran_pelatihan_industry': student.kehadiran_pelatihan_industry,
            'motivasi_akademik': student.motivasi_akademik,
            'exam_score': student.exam_score,
            'gender': student.gender,
            'rata_rata_pemasukan_keluarga': student.rata_rata_pemasukan_keluarga,
            'pendidikan_terakhir_orang_tua': student.pendidikan_terakhir_orang_tua,
            'kerja_sampingan': student.kerja_sampingan,
            'study_environment': student.study_environment,
            'kompetensi_skill_level': student.kompetensi_skill_level,
            'industry_readiness': student.industry_readiness,
            'stress_level': student.stress_level,
            'predicted_score': prediction.predicted_exam_score if prediction else None,
            'risk_status': prediction.risk_status if prediction else None,
            'latest_prediction': {
                'predicted_exam_score': prediction.predicted_exam_score if prediction else None,
                'risk_status': prediction.risk_status if prediction else None,
                'created_at': prediction.created_at if prediction else None,
            } if prediction else None,
        })

    return rows


def _analytics_summary(rows: list[dict[str, Any]]):
    total_students = len(rows)
    scores = [row['predicted_score'] for row in rows if isinstance(row.get('predicted_score'), (int, float))]
    avg_score = (sum(scores) / len(scores)) if scores else None

    study_hours = [row['jam_belajar_per_hari'] for row in rows if isinstance(row.get('jam_belajar_per_hari'), (int, float))]
    avg_study_hours = (sum(study_hours) / len(study_hours)) if study_hours else None

    risk_counts = {
        'Rendah': 0,
        'Netral': 0,
        'Tinggi': 0,
    }
    for row in rows:
        risk_status = row.get('risk_status')
        if risk_status in risk_counts:
            risk_counts[risk_status] += 1

    return {
        'total_students': total_students,
        'avg_score': avg_score,
        'avg_study_hours': avg_study_hours,
        'risk_counts': risk_counts,
        'predicted_students': len(scores),
    }


def _build_dashboard_card_table(title: str, value: str, caption: str):
    return Table([
        [Paragraph(f'<font color="#64748B" size="8"><b>{title}</b></font>', getSampleStyleSheet()['BodyText'])],
        [Paragraph(f'<font color="#0F172A" size="20"><b>{value}</b></font>', getSampleStyleSheet()['BodyText'])],
        [Paragraph(f'<font color="#64748B" size="8">{caption}</font>', getSampleStyleSheet()['BodyText'])],
    ], colWidths=[114])


def _build_dashboard_cards(summary: dict[str, Any]):
    rows = [[
        _build_dashboard_card_table('Total sample', str(summary['total_students']), 'Seluruh siswa di backend'),
        _build_dashboard_card_table('Avg exam score', '-' if summary['avg_score'] is None else f"{summary['avg_score']:.1f}", 'Rata-rata prediksi skor'),
        _build_dashboard_card_table('Avg study hours', '-' if summary['avg_study_hours'] is None else f"{summary['avg_study_hours']:.1f}", 'Jam belajar per hari'),
        _build_dashboard_card_table('Risk count', str(summary['risk_counts']['Tinggi']), 'Tinggi'),
    ]]
    table = Table(rows, colWidths=[122, 122, 122, 122])
    table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    return table


def _build_dashboard_hero():
    title_style = ParagraphStyle(
        name='AnalyticsHeroTitle',
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.white,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        name='AnalyticsHeroSubtitle',
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#DCFCE7'),
    )
    hero = Table([
        [Paragraph('ANALYTICS', ParagraphStyle(name='AnalyticsHeroKicker', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#DCFCE7'), leading=14))],
        [Paragraph('EDA Dashboard', title_style)],
        [Paragraph('Visualisasi eksploratif yang dibangun dari data siswa dan hasil analisis backend, bukan lagi mock data.', subtitle_style)],
    ], colWidths=[520])
    hero.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1F6F5F')),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#1F6F5F')),
        ('LEFTPADDING', (0, 0), (-1, -1), 20),
        ('RIGHTPADDING', (0, 0), (-1, -1), 20),
        ('TOPPADDING', (0, 0), (-1, -1), 18),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 18),
    ]))
    return hero


def _analytics_report_pdf():
    rows = _analytics_dashboard_rows()
    summary = _analytics_summary(rows)

    buffer = io.BytesIO()
    try:
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=24 * mm,
            bottomMargin=18 * mm,
        )
        styles = getSampleStyleSheet()
        styles.add(
            ParagraphStyle(
                name='SmallMuted',
                parent=styles['Normal'],
                fontSize=9,
                leading=12,
                textColor=colors.HexColor('#334155')
            )
        )
        story = []

        story.append(_build_dashboard_hero())
        story.append(Spacer(1, 12))
        story.append(_build_dashboard_cards(summary))
        story.append(Spacer(1, 12))

        chart_row = [[
            _build_risk_pie_chart(rows),
            _build_stress_bar_chart(rows),
        ], [
            _build_attendance_bar_chart(rows),
            Table([
                [Paragraph('<font color="#64748B" size="8"><b>Risk snapshot</b></font>', styles['BodyText'])],
                [Paragraph(f'<font color="#0F172A" size="18"><b>{summary["risk_counts"]["Rendah"]}</b></font>', styles['BodyText'])],
                [Paragraph('<font color="#64748B" size="8">Rendah</font>', styles['BodyText'])],
                [Spacer(1, 5)],
                [Paragraph(f'<font color="#0F172A" size="18"><b>{summary["risk_counts"]["Netral"]}</b></font>', styles['BodyText'])],
                [Paragraph('<font color="#64748B" size="8">Netral</font>', styles['BodyText'])],
                [Spacer(1, 5)],
                [Paragraph(f'<font color="#0F172A" size="18"><b>{summary["risk_counts"]["Tinggi"]}</b></font>', styles['BodyText'])],
                [Paragraph('<font color="#64748B" size="8">Tinggi</font>', styles['BodyText'])],
            ], colWidths=[250])
        ]]
        chart_table = Table(chart_row, colWidths=[260, 260])
        chart_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(chart_table)

        story.append(PageBreak())
        story.append(Paragraph('Siswa Prioritas', styles['Heading2']))
        story.append(Spacer(1, 8))
        story.append(
            Paragraph(
                'Lima siswa dengan risiko tertinggi ditampilkan sebagai tindak lanjut cepat untuk guru.',
                styles['SmallMuted']
            )
        )
        story.append(Spacer(1, 8))

        priority_rows = _build_top_risky_students_table(rows)
        priority_table = LongTable(priority_rows, repeatRows=1)
        priority_table.setStyle(_build_table_style())
        story.append(priority_table)

        doc.build(
            story,
            onFirstPage=_draw_vocavision_header_footer,
            onLaterPages=_draw_vocavision_header_footer,
        )
        buffer.seek(0)
        return buffer
    except Exception as e:
        # Fallback ultra-minimal: hindari ketergantungan SimpleDocTemplate (yang bisa gagal di kondisi ekstrem)
        print("[export/pdf] _analytics_report_pdf error:", repr(e))
        from reportlab.pdfgen import canvas as pdfcanvas
        from reportlab.lib.pagesizes import A4 as _A4

        fallback = io.BytesIO()
        c = pdfcanvas.Canvas(fallback, pagesize=landscape(_A4))
        c.setFont("Helvetica", 12)
        c.drawString(72, 720, "Gagal membuat PDF analytics.")
        c.setFont("Helvetica", 9)
        c.drawString(72, 700, f"Error: {repr(e)}")
        c.drawString(72, 680, "Silakan coba ekspor lagi atau cek data.")
        c.showPage()
        c.save()
        fallback.seek(0)
        return fallback



def _student_report_pdf(report: dict[str, Any]):
    buffer = io.BytesIO()
    try:
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=24 * mm,
            bottomMargin=18 * mm,
        )
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='ReportTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#0F172A')))
        styles.add(ParagraphStyle(name='ReportSubtitle', parent=styles['Normal'], fontSize=10, leading=14, textColor=colors.HexColor('#334155')))
        styles.add(ParagraphStyle(name='SmallMuted', parent=styles['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#334155')))
        styles.add(ParagraphStyle(name='BodyLead', parent=styles['BodyText'], fontSize=10, leading=15, textColor=colors.HexColor('#0F172A')))
        story = []

        student = report['student']
        prediction = report['prediction']
        insight = _build_student_story(report)
        title_style = ParagraphStyle(
            name='StudentHeroTitle',
            fontName='Helvetica-Bold',
            fontSize=24,
            leading=28,
            textColor=colors.white,
            spaceAfter=2,
        )
        subtitle_style = ParagraphStyle(
            name='StudentHeroSubtitle',
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#DCFCE7'),
        )

        story.append(Table([
            [Paragraph('REPORT AKADEMIK PER SISWA', ParagraphStyle(name='StudentHeroKicker', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#DCFCE7'), leading=14))],
            [Paragraph(student.nama_siswa, title_style)],
            [Paragraph(f'NISN {student.nisn} • Status {prediction.risk_status} • Prediksi {"-" if prediction.predicted_exam_score is None else f"{prediction.predicted_exam_score:.2f}"}', subtitle_style)],
        ], colWidths=[520], style=TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1F6F5F')),
            ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#1F6F5F')),
            ('LEFTPADDING', (0, 0), (-1, -1), 20),
            ('RIGHTPADDING', (0, 0), (-1, -1), 20),
            ('TOPPADDING', (0, 0), (-1, -1), 18),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 18),
        ])))
        story.append(Spacer(1, 12))

        summary_score = '-' if prediction.predicted_exam_score is None else f'{prediction.predicted_exam_score:.1f}'
        summary_attendance = '-' if student.presentase_kehadiran is None else f'{float(student.presentase_kehadiran):.0f}%'
        summary_study = '-' if student.jam_belajar_per_hari is None else f'{float(student.jam_belajar_per_hari):.1f} jam'

        metrics = Table([[
            Table([
                [Paragraph('<font color="#64748B" size="8"><b>Prediksi skor</b></font>', styles['BodyText'])],
                [Paragraph(f'<font color="#0F172A" size="18"><b>{summary_score}</b></font>', styles['BodyText'])],
                [Paragraph('<font color="#64748B" size="8">Hasil model terbaru</font>', styles['BodyText'])],
            ], colWidths=[114]),
            Table([
                [Paragraph('<font color="#64748B" size="8"><b>Status risiko</b></font>', styles['BodyText'])],
                [Paragraph(f'<font color="#0F172A" size="18"><b>{prediction.risk_status}</b></font>', styles['BodyText'])],
                [Paragraph('<font color="#64748B" size="8">Klasifikasi risiko terbaru</font>', styles['BodyText'])],
            ], colWidths=[114]),
            Table([
                [Paragraph('<font color="#64748B" size="8"><b>Kehadiran</b></font>', styles['BodyText'])],
                [Paragraph(f'<font color="#0F172A" size="18"><b>{summary_attendance}</b></font>', styles['BodyText'])],
                [Paragraph('<font color="#64748B" size="8">Persentase kehadiran</font>', styles['BodyText'])],
            ], colWidths=[114]),
            Table([
                [Paragraph('<font color="#64748B" size="8"><b>Jam belajar</b></font>', styles['BodyText'])],
                [Paragraph(f'<font color="#0F172A" size="18"><b>{summary_study}</b></font>', styles['BodyText'])],
                [Paragraph('<font color="#64748B" size="8">Jam belajar per hari</font>', styles['BodyText'])],
            ], colWidths=[114]),
        ]], colWidths=[122, 122, 122, 122])
        metrics.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(metrics)
        story.append(Spacer(1, 10))

        story.append(Paragraph(insight['summary_text'], styles['SmallMuted']))
        story.append(Spacer(1, 8))

        insight_table = Table([[
            Paragraph('<font color="#64748B" size="8"><b>Keunggulan</b></font>', styles['BodyText']),
            Paragraph('<font color="#64748B" size="8"><b>Perhatian</b></font>', styles['BodyText']),
            Paragraph('<font color="#64748B" size="8"><b>Tindak lanjut</b></font>', styles['BodyText']),
        ], [
            Paragraph('<br/>'.join([f'• {item}' for item in insight['strengths'][:3]]) or '• Belum ada poin yang tercatat.', styles['BodyText']),
            Paragraph('<br/>'.join([f'• {item}' for item in insight['concerns'][:3]]) or '• Belum ada poin yang tercatat.', styles['BodyText']),
            Paragraph('<br/>'.join([f'• {item}' for item in insight['actions'][:3]]) or '• Belum ada poin yang tercatat.', styles['BodyText']),
        ]], colWidths=[170, 170, 170])
        insight_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('BOX', (0, 1), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ]))
        story.append(insight_table)
        story.append(Spacer(1, 10))

        if insight['top_shap_texts']:
            story.append(Paragraph('Faktor SHAP Utama', styles['BodyLead']))
            story.append(Spacer(1, 4))
            shap_lines = ''.join([f'<br/>• {text}' for text in insight['top_shap_texts']])
            story.append(Paragraph(shap_lines, styles['SmallMuted']))
            story.append(Spacer(1, 8))

        intervention_rows = _student_intervention_rows(report)
        intervention_table = LongTable(intervention_rows[:4], repeatRows=1)
        intervention_table.setStyle(_build_table_style())
        story.append(Paragraph('Catatan Tindak Lanjut', styles['BodyLead']))
        story.append(Spacer(1, 4))
        story.append(intervention_table)

        doc.build(
            story,
            onFirstPage=_draw_vocavision_header_footer,
            onLaterPages=_draw_vocavision_header_footer,
        )
        buffer.seek(0)
        return buffer
    except Exception as e:
        # Fallback ultra-minimal: pastikan buffer terisi meskipun SimpleDocTemplate gagal
        print("[export/pdf] _student_report_pdf error:", repr(e))
        from reportlab.pdfgen import canvas as pdfcanvas
        from reportlab.lib.pagesizes import A4 as _A4

        fallback = io.BytesIO()
        c = pdfcanvas.Canvas(fallback, pagesize=landscape(_A4))
        c.setFont("Helvetica", 12)
        c.drawString(72, 720, "Gagal membuat PDF laporan siswa.")
        c.setFont("Helvetica", 9)
        c.drawString(72, 700, f"Error: {repr(e)}")
        c.drawString(72, 680, "Silakan coba ekspor lagi atau periksa data siswa.")
        c.showPage()
        c.save()
        fallback.seek(0)
        if len(fallback.getvalue()) <= 0:
            print("[export/pdf] _student_report_pdf fallback buffer still empty")
        return fallback



def _student_report_excel(report: dict[str, Any]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = 'Profil'
    ws_summary.append(['Field', 'Value'])
    for title, value in _student_summary_rows(report):
        ws_summary.append([title, value])

    ws_variables = wb.create_sheet('Data Siswa')
    for row in _student_variable_rows(report):
        ws_variables.append(row)

    ws_shap = wb.create_sheet('Faktor Prediksi')
    for row in _student_shap_rows(report):
        ws_shap.append(row)

    ws_interventions = wb.create_sheet('Intervensi Guru')
    for row in _student_intervention_rows(report):
        ws_interventions.append(row)

    header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for worksheet in wb.worksheets:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for column_index, _ in enumerate(worksheet[1], start=1):
            max_len = len(str(worksheet.cell(row=1, column=column_index).value or ''))
            for row_index in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 12), 42)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@export_bp.route('/pdf', methods=['GET'])
@jwt_required()
def export_pdf():
    try:
        student_id = request.args.get('student_id')
        if student_id:
            report, error = _build_student_report(student_id)
            if error:
                message, status_code = error
                print("[export/pdf] student export build error:", student_id, status_code, message)
                return jsonify({'message': message}), status_code

            buffer = _student_report_pdf(report)
            buffer_bytes = len(buffer.getvalue()) if buffer else 0
            print("[export/pdf] student export result:", student_id, "bytes=", buffer_bytes)

            if buffer_bytes <= 0:
                # Prevent returning an empty PDF with HTTP 200
                return jsonify({
                    'message': 'PDF buffer is empty',
                    'student_id': student_id,
                }), 500

            student_slug = _slugify_filename(report['student'].nama_siswa)
            filename = f'laporan-{student_slug}-{student_id}.pdf'
        else:
            buffer = _analytics_report_pdf()
            buffer_bytes = len(buffer.getvalue()) if buffer else 0
            if buffer_bytes <= 0:
                return jsonify({
                    'message': 'Analytics PDF buffer is empty',
                }), 500

            filename = 'laporan-analytics-guru.pdf'

        pdf_bytes = buffer.getvalue() if buffer else b""
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Length'] = str(len(pdf_bytes))
        return response
    except Exception as e:
        print("[export/pdf] top-level error:", repr(e))
        return jsonify({'message': 'Gagal membuat PDF', 'error': str(e)}), 500



@export_bp.route('/excel', methods=['GET'])
@jwt_required()
def export_excel():
    try:
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({'message': 'student_id diperlukan'}), 400

        report, error = _build_student_report(student_id)
        if error:
            message, status_code = error
            return jsonify({'message': message}), status_code

        buffer = _student_report_excel(report)
        student_slug = _slugify_filename(report['student'].nama_siswa)
        filename = f'laporan-{student_slug}-{student_id}.xlsx'

        excel_bytes = buffer.getvalue() if buffer else b""
        response = make_response(excel_bytes)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Length'] = str(len(excel_bytes))
        return response
    except Exception as e:
        return jsonify({'message': 'Gagal membuat Excel', 'error': str(e)}), 500
